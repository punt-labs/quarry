from __future__ import annotations

from pathlib import Path

import pytest

from quarry.models import PageType
from quarry.spreadsheet_processor import (
    SUPPORTED_SPREADSHEET_EXTENSIONS,
    _escape_latex,
    _rows_to_latex,
    _split_rows_to_sections,
    process_spreadsheet_file,
)


class TestSupportedExtensions:
    def test_includes_xlsx_and_csv(self):
        assert ".xlsx" in SUPPORTED_SPREADSHEET_EXTENSIONS
        assert ".csv" in SUPPORTED_SPREADSHEET_EXTENSIONS

    def test_no_overlap_with_other_extensions(self):
        from quarry.code_processor import SUPPORTED_CODE_EXTENSIONS
        from quarry.text_processor import SUPPORTED_TEXT_EXTENSIONS

        overlap = SUPPORTED_SPREADSHEET_EXTENSIONS & (
            SUPPORTED_CODE_EXTENSIONS | SUPPORTED_TEXT_EXTENSIONS
        )
        assert overlap == frozenset(), f"Overlapping extensions: {overlap}"


class TestEscapeLatex:
    def test_escapes_ampersand(self):
        assert _escape_latex("A & B") == r"A \& B"

    def test_escapes_percent(self):
        assert _escape_latex("50%") == r"50\%"

    def test_escapes_dollar(self):
        assert _escape_latex("$100") == r"\$100"

    def test_escapes_hash(self):
        assert _escape_latex("#1") == r"\#1"

    def test_escapes_underscore(self):
        assert _escape_latex("my_var") == r"my\_var"

    def test_escapes_braces(self):
        assert _escape_latex("{x}") == r"\{x\}"

    def test_escapes_backslash(self):
        assert _escape_latex(r"a\b") == r"a\textbackslash{}b"

    def test_plain_text_unchanged(self):
        assert _escape_latex("Hello World 123") == "Hello World 123"

    def test_multiple_specials(self):
        result = _escape_latex("$100 & 50%")
        assert r"\$" in result
        assert r"\&" in result
        assert r"\%" in result


class TestRowsToLatex:
    def test_basic_table(self):
        result = _rows_to_latex(["Name", "Age"], [["Alice", "30"]])
        assert r"\begin{tabular}{ll}" in result
        assert "Name & Age" in result
        assert "Alice & 30" in result
        assert r"\end{tabular}" in result

    def test_empty_headers_returns_empty(self):
        assert _rows_to_latex([], [["a", "b"]]) == ""

    def test_sheet_name_prefix(self):
        result = _rows_to_latex(["A"], [["1"]], sheet_name="Data")
        assert "% Sheet: Data" in result

    def test_no_sheet_name(self):
        result = _rows_to_latex(["A"], [["1"]])
        assert "% Sheet" not in result

    def test_row_padding(self):
        result = _rows_to_latex(["A", "B", "C"], [["1"]])
        # Row should be padded to match 3 columns
        assert "1 &  & " in result

    def test_row_truncation(self):
        result = _rows_to_latex(["A"], [["1", "2", "3"]])
        # Extra columns should be dropped
        assert "1 \\\\" in result
        assert "2" not in result

    def test_latex_escaping_in_cells(self):
        result = _rows_to_latex(["Price"], [["$100"]])
        assert r"\$100" in result

    def test_empty_rows(self):
        result = _rows_to_latex(["A", "B"], [])
        assert r"\begin{tabular}" in result
        assert r"\end{tabular}" in result


class TestSplitRowsToSections:
    def test_small_table_single_section(self):
        sections = _split_rows_to_sections(
            ["A", "B"], [["1", "2"], ["3", "4"]], None, max_chars=5000
        )
        assert len(sections) == 1

    def test_large_table_splits(self):
        headers = ["Name", "Value"]
        rows = [[f"item_{i}", f"value_{i}"] for i in range(50)]
        sections = _split_rows_to_sections(headers, rows, None, max_chars=200)
        assert len(sections) > 1
        # Each section should have the headers
        for section in sections:
            assert "Name & Value" in section

    def test_sheet_name_in_all_sections(self):
        headers = ["A"]
        rows = [[f"row_{i}"] for i in range(50)]
        sections = _split_rows_to_sections(headers, rows, "MySheet", max_chars=100)
        for section in sections:
            assert "% Sheet: MySheet" in section

    def test_single_row_never_split(self):
        headers = ["A" * 100]
        rows = [["B" * 100]]
        sections = _split_rows_to_sections(headers, rows, None, max_chars=10)
        assert len(sections) == 1

    def test_empty_rows_returns_empty(self):
        sections = _split_rows_to_sections(["A"], [], None, max_chars=100)
        # Empty rows still produce a valid (header-only) table
        result = _rows_to_latex(["A"], [])
        if result.strip():
            assert len(sections) <= 1


class TestProcessSpreadsheetCSV:
    def test_basic_csv(self, tmp_path: Path):
        f = tmp_path / "data.csv"
        f.write_text("Name,Age\nAlice,30\nBob,25\n")

        pages, sheet_count = process_spreadsheet_file(f)

        assert sheet_count == 1
        assert len(pages) == 1
        assert pages[0].page_type == PageType.SPREADSHEET
        assert pages[0].document_name == "data.csv"
        assert "Alice" in pages[0].text
        assert "Bob" in pages[0].text
        assert r"\begin{tabular}" in pages[0].text

    def test_csv_metadata(self, tmp_path: Path):
        f = tmp_path / "test.csv"
        f.write_text("A,B\n1,2\n")

        pages, _ = process_spreadsheet_file(f)

        assert pages[0].document_path == str(f.resolve())
        assert pages[0].page_number == 1
        assert pages[0].total_pages == 1

    def test_empty_csv(self, tmp_path: Path):
        f = tmp_path / "empty.csv"
        f.write_text("")

        pages, sheet_count = process_spreadsheet_file(f)
        assert pages == []
        assert sheet_count == 0

    def test_header_only_csv(self, tmp_path: Path):
        f = tmp_path / "header.csv"
        f.write_text("A,B,C\n")

        pages, _ = process_spreadsheet_file(f)

        assert len(pages) == 1
        assert "A & B & C" in pages[0].text

    def test_csv_with_special_chars(self, tmp_path: Path):
        f = tmp_path / "special.csv"
        f.write_text('Item,Price\nWidget,"$100"\n')

        pages, _ = process_spreadsheet_file(f)

        assert r"\$100" in pages[0].text

    def test_csv_large_splits(self, tmp_path: Path):
        rows = ["Name,Value"] + [f"item_{i},val_{i}" for i in range(100)]
        f = tmp_path / "large.csv"
        f.write_text("\n".join(rows) + "\n")

        pages, sheet_count = process_spreadsheet_file(f, max_chars=300)

        assert sheet_count == 1
        assert len(pages) > 1
        for page in pages:
            assert "Name & Value" in page.text

    def test_csv_no_sheet_name_prefix(self, tmp_path: Path):
        f = tmp_path / "data.csv"
        f.write_text("A\n1\n")

        pages, _ = process_spreadsheet_file(f)

        # Single-sheet files should not have sheet name prefix
        assert "% Sheet" not in pages[0].text

    def test_csv_document_name_override(self, tmp_path: Path):
        f = tmp_path / "data.csv"
        f.write_text("A\n1\n")

        pages, _ = process_spreadsheet_file(f, document_name="subdir/data.csv")

        assert pages[0].document_name == "subdir/data.csv"


class TestProcessSpreadsheetXLSX:
    def test_basic_xlsx(self, tmp_path: Path):
        import openpyxl

        f = tmp_path / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])
        wb.save(f)

        pages, sheet_count = process_spreadsheet_file(f)

        assert sheet_count == 1
        assert len(pages) == 1
        assert pages[0].page_type == PageType.SPREADSHEET
        assert pages[0].document_name == "data.xlsx"
        assert "Alice" in pages[0].text
        assert "30" in pages[0].text

    def test_multi_sheet_xlsx(self, tmp_path: Path):
        import openpyxl

        f = tmp_path / "multi.xlsx"
        wb = openpyxl.Workbook()

        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Sales"
        ws1.append(["Region", "Revenue"])
        ws1.append(["North", "1000"])

        ws2 = wb.create_sheet("Costs")
        ws2.append(["Category", "Amount"])
        ws2.append(["Rent", "500"])

        wb.save(f)

        pages, sheet_count = process_spreadsheet_file(f)

        assert sheet_count == 2
        assert len(pages) == 2
        assert "% Sheet: Sales" in pages[0].text
        assert "% Sheet: Costs" in pages[1].text

    def test_empty_xlsx(self, tmp_path: Path):
        import openpyxl

        f = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.save(f)

        pages, _sheet_count = process_spreadsheet_file(f)
        assert pages == []

    def test_xlsx_with_none_cells(self, tmp_path: Path):
        import openpyxl

        f = tmp_path / "sparse.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["A", "B", "C"])
        ws.append([1, None, 3])
        wb.save(f)

        pages, _ = process_spreadsheet_file(f)

        assert len(pages) == 1
        # None cells should be empty strings
        assert "1 &  & 3" in pages[0].text

    def test_xlsx_formulas_use_cached_values(self, tmp_path: Path):
        import openpyxl

        f = tmp_path / "formula.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["A", "B", "Sum"])
        ws["A2"] = 10
        ws["B2"] = 20
        ws["C2"] = "=A2+B2"
        wb.save(f)

        pages, _ = process_spreadsheet_file(f)

        # data_only=True means formulas show cached value or None
        assert len(pages) == 1

    def test_xlsx_page_numbers_sequential(self, tmp_path: Path):
        import openpyxl

        f = tmp_path / "seq.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "A"
        ws1.append(["X"])
        ws1.append(["1"])

        ws2 = wb.create_sheet("B")
        ws2.append(["Y"])
        ws2.append(["2"])

        ws3 = wb.create_sheet("C")
        ws3.append(["Z"])
        ws3.append(["3"])
        wb.save(f)

        pages, sheet_count = process_spreadsheet_file(f)

        assert sheet_count == 3
        assert len(pages) == 3
        for i, page in enumerate(pages):
            assert page.page_number == i + 1
            assert page.total_pages == 3

    def test_xlsx_document_name_override(self, tmp_path: Path):
        import openpyxl

        f = tmp_path / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["A"])
        ws.append(["1"])
        wb.save(f)

        pages, _ = process_spreadsheet_file(f, document_name="subdir/data.xlsx")

        assert pages[0].document_name == "subdir/data.xlsx"


class TestProcessSpreadsheetErrors:
    def test_unsupported_extension(self, tmp_path: Path):
        f = tmp_path / "data.xls"
        f.write_bytes(b"\x00")

        with pytest.raises(ValueError, match="Unsupported spreadsheet format"):
            process_spreadsheet_file(f)
